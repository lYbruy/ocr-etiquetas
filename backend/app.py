from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from pydantic import BaseModel

import os
import re
import cv2
import uuid
import csv
import shutil
import traceback
import requests
import numpy as np
from concurrent.futures import ThreadPoolExecutor

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from paddleocr import PaddleOCR


# =========================
# CONFIG
# =========================

PREFIXOS_ACEITES = ["3800", "3810"]

EXPORT_EXCEL = "exports/resultado.xlsx"
EXPORT_CSV = "exports/resultado.csv"

GEOCODER_ENABLED = os.getenv("GEOCODER_ENABLED", "true").lower() == "true"
GEOCODER_TIMEOUT = float(os.getenv("GEOCODER_TIMEOUT", "1.5"))
GEOCODER_STRICT = os.getenv("GEOCODER_STRICT", "false").lower() == "true"
GEOCODER_USER_AGENT = os.getenv(
    "GEOCODER_USER_AGENT",
    "ocr-etiquetas-aveiro/1.0"
)

LOCALIDADES_AVEIRO = [
    "AVEIRO", "CACIA", "CÁCIA", "ESGUEIRA", "ARADAS",
    "GLORIA", "GLÓRIA", "VERA CRUZ", "SANTA JOANA",
    "SAO BERNARDO", "SÃO BERNARDO", "OLIVEIRINHA",
    "EIXO", "EIROL", "NARIZ", "REQUEIXO",
    "NOSSA SENHORA DE FATIMA", "NOSSA SENHORA DE FÁTIMA",
]

PALAVRAS_NAO_PORTUGAL = [
    "SPAIN", "ESPAÑA", "ESPANA", "ESPANHA",
    "CALLE ", "CALLE,", " C/ ", "C/.", " CL ",
    "PLAZA ", "PASEO ", "PASE0 ", "CARRER ",
    "AVDA.", "AVDA ", "POL. ", "POL,",
    "POLIGONO", "POLÍGONO",
    "NAVE ", "PARCELA ",
    "P.O. BOX", "APARTADO DE CORREOS",
    "GERMANY", "FRANCE", "ITALIA",
    "UNITED KINGDOM", "NEDERLAND", "BELGIQUE",
]

CIDADES_NAO_PORTUGAL = [
    "MADRID", "BARCELONA", "VALENCIA", "SEVILLA", "ZARAGOZA",
    "MALAGA", "MURCIA", "BILBAO", "ALICANTE", "CORDOBA",
    "VALLADOLID", "VIGO", "GIJON", "VITORIA", "GRANADA",
    "ELCHE", "OVIEDO", "BADALONA", "CARTAGENA", "TERRASSA",
    "SABADELL", "JEREZ", "MOSTOLES", "LEGANES", "BURGOS",
    "SANTANDER", "FUENLABRADA", "ALMERIA", "ALCALA", "PAMPLONA",
    "CADIZ", "SALAMANCA", "TOLEDO", "HUELVA", "BADAJOZ",
    "LOGRONO", "TARRAGONA", "LLEIDA", "GIRONA", "ALBACETE",
    "BERLIN", "PARIS", "LONDON", "ROMA", "AMSTERDAM",
    "BRUSSELS", "LISBOA",
]

PREFIXOS_CP_ESPANHA = set(f"{n:02d}" for n in range(1, 53))


# =========================
# FASTAPI
# =========================

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================
# PASTAS
# =========================

os.makedirs("uploads", exist_ok=True)
os.makedirs("exports", exist_ok=True)


# =========================
# MEMÓRIA DO LOTE
# =========================

ocr_engine = None
uploads_pendentes = {}
lote_confirmado = []
executor = ThreadPoolExecutor(max_workers=3)


# =========================
# MODELOS
# =========================

class ConfirmarPayload(BaseModel):
    upload_id: str | None = None
    morada: str
    codigo_postal: str
    cidade: str | None = ""
    texto_ocr: str | None = ""


# =========================
# OCR GLOBAL
# =========================

def get_ocr():
    global ocr_engine

    if ocr_engine is None:
        print("Inicializando PaddleOCR...", flush=True)

        ocr_engine = PaddleOCR(
            use_angle_cls=True,
            lang="en",
            show_log=False,
            det_db_thresh=0.3,
            det_db_box_thresh=0.5,
            rec_batch_num=6,
        )

        print("PaddleOCR inicializado", flush=True)

    return ocr_engine


@app.on_event("startup")
async def startup_event():
    try:
        limpar_pasta_exports()
        uploads_pendentes.clear()

        print("Pré-aquecendo OCR...", flush=True)

        engine = get_ocr()

        dummy = np.ones((100, 300, 3), dtype=np.uint8) * 255
        dummy_path = "uploads/_warmup.jpg"

        cv2.imwrite(dummy_path, dummy)
        engine.ocr(dummy_path, cls=False)

        if os.path.exists(dummy_path):
            os.remove(dummy_path)

        print("OCR pré-aquecido com sucesso!", flush=True)

    except Exception:
        traceback.print_exc()


# =========================
# ROTAS TESTE
# =========================

@app.get("/")
async def home():
    return {
        "status": "online",
        "message": "API OCR funcionando",
        "filtro": "Somente códigos postais 3800 e 3810 — moradas Portugal/Aveiro",
        "geocoder_enabled": GEOCODER_ENABLED,
        "geocoder_timeout": GEOCODER_TIMEOUT,
        "geocoder_strict": GEOCODER_STRICT,
    }


@app.get("/health")
async def health():
    return {"status": "ok"}


# =========================
# NORMALIZAÇÃO
# =========================

def limpar_linha(linha: str) -> str:
    linha = str(linha).strip()
    linha = linha.replace("º", "°")
    linha = linha.replace(" N ", " Nº ")
    linha = linha.replace(" N°", " Nº")
    linha = linha.replace(" N.", " Nº")
    linha = linha.replace(" N0 ", " Nº ")
    linha = linha.replace(" No ", " Nº ")
    linha = linha.replace("N°", "Nº")
    linha = linha.replace("N.", "Nº")
    linha = re.sub(r"\s+", " ", linha)
    return linha.strip()


def normalizar_texto(texto: str) -> str:
    texto = str(texto).replace("\r", "\n")
    texto = re.sub(r"\n+", "\n", texto)

    linhas = []

    for linha in texto.split("\n"):
        linha = limpar_linha(linha)

        if linha:
            linhas.append(linha)

    return "\n".join(linhas)


def corrigir_ocr_para_morada(texto: str) -> str:
    texto = str(texto)

    trocas = {
        "AVEIR0": "AVEIRO",
        "AYEIR0": "AVEIRO",
        "AVElRO": "AVEIRO",
        "AVElR0": "AVEIRO",
        "AVFIR0": "AVEIRO",
        "AVR0": "AVEIRO",
        "AVRO": "AVEIRO",
        "AVR": "AVEIRO",
        "AVARO": "AVEIRO",
        "AYARO": "AVEIRO",
        "AVERO": "AVEIRO",
        "AVE1R0": "AVEIRO",
        "AVE1RO": "AVEIRO",
        "E56UEIRA": "ESGUEIRA",
        "ESGUERA": "ESGUEIRA",
        "ESGUSA": "ESGUEIRA",
        "ESGUELSA": "ESGUEIRA",
        "PORTUGA": "PORTUGAL",
        "POR TUGAL": "PORTUGAL",
        "PORTUGALO.C": "PORTUGAL O.C",
        "PORTUGALO.O": "PORTUGAL O.C",
        "CACIA PORTUGALO.C.": "CACIA PORTUGAL",
        "CACIA PORTUGAL-O.C.": "CACIA PORTUGAL",
        "A1AMEDA": "ALAMEDA",
        "A1ameda": "Alameda",
        "S1LVA": "SILVA",
        "Si1va": "Silva",
        "R0CHA": "ROCHA",
        "R0A": "RUA",
        "RU4": "RUA",
        "REPOBLICA": "REPUBLICA",
        "REP0BLICA": "REPUBLICA",
        "REPÚBLICA": "REPUBLICA",
        "NACIONAD": "NACIONAL",
        "NACLONA": "NACIONAL",
        "NACIONA": "NACIONAL",
        "EST.NAC": "ESTRADA NACIONAL",
        "EST NAC": "ESTRADA NACIONAL",
        "EUROPAN": "EUROPA Nº",
        "EUROPA N": "EUROPA Nº",
        "AVENIDAEUROPA": "AVENIDA EUROPA",
        "AVENIDA EUROPA N292": "AVENIDA EUROPA Nº292",
        "AVENIDA EUROPA N°292": "AVENIDA EUROPA Nº292",
        "AVENIDA EUROPA Nº 292": "AVENIDA EUROPA Nº292",
    }

    for errado, certo in trocas.items():
        texto = texto.replace(errado, certo)

    texto = re.sub(
        r"\b(AVENIDA|RUA|ALAMEDA|TRAVESSA|ESTRADA|CAMINHO|LARGO|PRAÇA|PRACA|PRACETA)\s*([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]+?)N[º°]?\s*(\d+)",
        r"\1 \2 Nº\3",
        texto,
        flags=re.IGNORECASE,
    )

    texto = re.sub(r"\s+", " ", texto)

    return limpar_linha(texto)


def converter_ocr_numero(texto: str) -> str:
    texto = str(texto).upper()
    texto = texto.replace("O", "0").replace("Q", "0").replace("D", "0")
    texto = texto.replace("I", "1").replace("L", "1").replace("|", "1")
    texto = texto.replace("S", "5").replace("B", "8").replace("G", "6")
    return texto


def normalizar_morada_extraida(morada: str) -> str:
    m = corrigir_ocr_para_morada(str(morada)).upper()

    trocas = {
        " Nº ": " Nº ",
        " N° ": " Nº ",
        " N. ": " Nº ",
        " NR ": " Nº ",
        " NR": " Nº",
        " N ": " Nº ",
        " NO ": " Nº ",

        "REP0BLICA": "REPUBLICA",
        "REPOBLICA": "REPUBLICA",
        "REPÚBLICA": "REPUBLICA",
        "RUADAREPUBLICA": "RUA DA REPUBLICA",
        "RUA DAREPUBLICA": "RUA DA REPUBLICA",
        "RUA DAREP0BLICA": "RUA DA REPUBLICA",
        "RUA DAREPOBLICA": "RUA DA REPUBLICA",
        "DAREPUBLICA": "DA REPUBLICA",
        "DAREPOBLICA": "DA REPUBLICA",

        "DOUT0R": "DOUTOR",
        "D0UTOR": "DOUTOR",
        "DR ": "DOUTOR ",
        "JOSE": "JOSÉ",
        "GRACA": "GRAÇA",

        "PORTAO": "PORTÃO",
        "PORTA0": "PORTÃO",
        "VIVIENDA": "VIVENDA",

        "TRAVESSADO": "TRAVESSA DO",
        "TRAVESSAD0": "TRAVESSA DO",
        "TRAVESSA D0": "TRAVESSA DO",
        "TRAVESSA0": "TRAVESSA DO",
        "TRAVESSA D ": "TRAVESSA DO ",
        "MILAO": "MILÃO",
        "MIL4O": "MILÃO",
        "M1LAO": "MILÃO",

        "RUADAPAZ": "RUA DA PAZ",
        "RUA DAPAZ": "RUA DA PAZ",
        "RUA DA PA2": "RUA DA PAZ",
        "CAC1A": "CACIA",
        "CÁCIA": "CACIA",
        "CACA": "CACIA",

        "PCT DARUADA": "PCT DA RUA DA",
        "PCT DA RUA": "PRACETA DA RUA",
        "PCT ": "PRACETA ",

        "SANTARI0S": "SANITÁRIOS",
        "SANTARIOS": "SANITÁRIOS",
        "SANITARIOS": "SANITÁRIOS",
        "SIOML": "SEIXAL",
        "SEI XAL": "SEIXAL",
        "SEIX0L": "SEIXAL",

        "AVEIR0": "AVEIRO",
        "AVE1RO": "AVEIRO",
        "AVE1R0": "AVEIRO",
        "AVERO": "AVEIRO",
        "AVARO": "AVEIRO",
        "AYARO": "AVEIRO",
        "ESGUERA": "ESGUEIRA",
        "E56UEIRA": "ESGUEIRA",
        "ESGUSA": "ESGUEIRA",
        "ESGUELSA": "ESGUEIRA",
    }

    for errado, certo in trocas.items():
        m = m.replace(errado, certo)

    m = re.sub(r"\bNº\s*O\b", "Nº 0", m)

    m = re.sub(r"([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ])(\d)", r"\1 \2", m)
    m = re.sub(r"(\d)([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ])", r"\1 \2", m)

    m = re.sub(r"\bN\s*[º°]?\s*(\d+)", r"Nº \1", m)
    m = re.sub(r"\bNR\s*(\d+)", r"Nº \1", m)

    m = m.replace(" ,", ",")
    m = re.sub(r"[|_]+", " ", m)
    m = re.sub(r"\s+", " ", m).strip()

    return limpar_linha(m)


def _corrigir_tokens_ocr_morada(texto: str) -> str:
    """
    Corrige tokens curtos de OCR apenas contra vocabulário de morada.
    Não tenta adivinhar nomes completos: serve para palavras estruturais
    como RUA/AVENIDA/INDUSTRIAL/LOTE e localidades aceites.
    """
    import difflib

    texto = str(texto).upper()
    texto = re.sub(r"\bLNOUSIR.?AL\b", "INDUSTRIAL", texto)
    texto = re.sub(r"\b([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]{4,})NR\s*(\d+)", r"\1 NR \2", texto)
    texto = re.sub(r"\bTABOEIRATAB\b", "TABOEIRA TAB", texto)

    vocabulario = set(PALAVRAS_FORMATO_MORADA + LOCALIDADES_AVEIRO)
    tokens = re.split(r"(\W+)", texto)
    corrigidos = []

    for token in tokens:
        if (
            len(token) < 5
            or not re.search(r"[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]", token)
            or re.search(r"\d", token)
        ):
            corrigidos.append(token)
            continue

        melhor = ""
        melhor_score = 0.0

        for palavra in vocabulario:
            if abs(len(palavra) - len(token)) > 3:
                continue

            score = difflib.SequenceMatcher(None, token, palavra).ratio()

            if score > melhor_score:
                melhor = palavra
                melhor_score = score

        if melhor and melhor_score >= 0.74:
            corrigidos.append(melhor)
        else:
            corrigidos.append(token)

    return "".join(corrigidos)


def _separar_palavras_coladas_morada(texto: str) -> str:
    t = str(texto).upper()

    artigos = ["DA", "DO", "DE", "DAS", "DOS"]
    vias = [
        "RUA", "AVENIDA", "ALAMEDA", "TRAVESSA", "LARGO",
        "PRACETA", "PRACA", "PRAÇA", "ESTRADA", "CAMINHO",
        "ZONA", "QUINTA", "LUGAR", "CASAL", "BAIRRO",
    ]

    for via in sorted(vias, key=len, reverse=True):
        t = re.sub(rf"\b{via}(?=[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ])", f"{via} ", t)

    for artigo in sorted(artigos, key=len, reverse=True):
        t = re.sub(
            rf"\b({'|'.join(vias)})\s*{artigo}(?=[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ])",
            rf"\1 {artigo} ",
            t,
        )
        t = re.sub(
            rf"([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]{{4,}}){artigo}([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]{{3,}})",
            rf"\1 {artigo} \2",
            t,
        )

    pares_comuns = [
        ("ZONA", "INDUSTRIAL"),
        ("INDUSTRIAL", "DE"),
        ("INDUSTRIAL", "DA"),
        ("INDUSTRIAL", "DO"),
        ("INDUSTRIAL", "DAS"),
        ("INDUSTRIAL", "DOS"),
        ("TABOEIRA", "TAB"),
        ("REPUBLICA", "NR"),
        ("REPUBLICA", "N"),
        ("FRANCISCO", "DO"),
        ("FRANCISCO", "DA"),
        ("JOAO", "FRANCISCO"),
    ]

    for antes, depois in pares_comuns:
        t = re.sub(rf"\b{antes}{depois}\b", f"{antes} {depois}", t)

    for artigo in sorted(artigos, key=len, reverse=True):
        t = re.sub(
            rf"\b([A-ZÃÃ€Ã‚ÃƒÃ‰ÃˆÃŠÃÃŒÃ“Ã’Ã”Ã•ÃšÃ™Ã‡]{{4,}}){artigo}\b",
            rf"\1 {artigo}",
            t,
        )
        t = re.sub(
            rf"\b{artigo}([A-ZÃÃ€Ã‚ÃƒÃ‰ÃˆÃŠÃÃŒÃ“Ã’Ã”Ã•ÃšÃ™Ã‡]{{4,}})\b",
            rf"{artigo} \1",
            t,
        )

    termos_estruturais = [
        "INDUSTRIAL", "NACIONAL", "REPUBLICA", "EUROPA", "TABOEIRA",
        "AVEIRO", "ESGUEIRA", "CACIA", "ARADAS", "LOTE", "BLOCO",
        "PORTA", "PISO", "ANDAR", "ARMAZEM", "ARMAZÃ‰M",
    ]

    for termo in sorted(termos_estruturais, key=len, reverse=True):
        t = re.sub(
            rf"\b([A-ZÃÃ€Ã‚ÃƒÃ‰ÃˆÃŠÃÃŒÃ“Ã’Ã”Ã•ÃšÃ™Ã‡]{{4,}}){termo}\b",
            rf"\1 {termo}",
            t,
        )
        t = re.sub(
            rf"\b{termo}([A-ZÃÃ€Ã‚ÃƒÃ‰ÃˆÃŠÃÃŒÃ“Ã’Ã”Ã•ÃšÃ™Ã‡]{{4,}})\b",
            rf"{termo} \1",
            t,
        )

    t = re.sub(r"\bAVENI\s+DA\b", "AVENIDA", t)
    t = re.sub(r"\bALAME\s+DA\b", "ALAMEDA", t)

    t = re.sub(r"\bLT\b", "LOTE", t)
    t = re.sub(r"([A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]{4,})NR\s*(\d+)", r"\1 Nº \2", t)
    t = re.sub(r"\b(REPUBLICA|REPÚBLICA)\s*N\s*R?\s*(\d+)", r"\1 Nº \2", t)
    t = re.sub(r"\bNR\s*(\d+)", r"Nº \1", t)
    t = re.sub(r"\bN\s*[º°]?\s*(\d+)", r"Nº \1", t)

    return re.sub(r"\s+", " ", t).strip()


def _remover_lixo_etiqueta_morada(texto: str) -> str:
    t = str(texto)

    for palavra in sorted(PALAVRAS_CORTE_ETIQUETA, key=len, reverse=True):
        padrao = r"(?i)(^|[\s,;|:/-])" + re.escape(palavra) + r"(\b|[:.])"
        m = re.search(padrao, t)

        if m:
            t = t[:m.start()].strip()

    t = re.sub(r"\b(3800|3810)\s*[- ]?\s*\d{3}\b.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(3800|3810)\b\s*$", "", t, flags=re.IGNORECASE)

    # Telefones, tracking e códigos internos longos que aparecem depois/colados à morada.
    t = re.sub(r"(?:\+?\s*351\s*)?9(?:\s*\d){8}\b.*$", "", t)
    t = re.sub(r"\b9\d{8}\b.*$", "", t)
    t = re.sub(r"\b\d{6,}\b.*$", "", t)
    t = re.sub(r"\b[A-Z]{1,6}\d{6,}[A-Z0-9]*\b.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b\d{2,}[A-Z]{2,}\d{3,}[A-Z0-9]*\b.*$", "", t, flags=re.IGNORECASE)

    return t.strip()


def _normalizar_numero_porta_morada(texto: str) -> str:
    t = str(texto)

    t = re.sub(r"\bN\s*[ÂºÂ°º]?\s*(\d{1,5}[A-Z]?)\b", r"NÂº \1", t, flags=re.IGNORECASE)
    t = re.sub(r"\bNR\s*(\d{1,5}[A-Z]?)\b", r"NÂº \1", t, flags=re.IGNORECASE)
    t = re.sub(r"\bNÂº\s*(\d{1,5})(ESQ|DTO|DIR|DRT|FTE|FRENTE|TRAS|TRASEIRAS)\b", r"NÂº \1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(\d{1,5})(ESQ|DTO|DIR|DRT|FTE|FRENTE|TRAS|TRASEIRAS)\b", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\bR\s*/\s*C\b", "R/C", t, flags=re.IGNORECASE)

    return re.sub(r"\s+", " ", t).strip()


def _normalizar_mojibake_saida(texto: str) -> str:
    t = str(texto)
    trocas = {
        "NÂº": "N\u00ba",
        "NÂ°": "N\u00ba",
        "Ã": "\u00c1",
        "Ã€": "\u00c0",
        "Ã‚": "\u00c2",
        "Ãƒ": "\u00c3",
        "Ã‰": "\u00c9",
        "Ãˆ": "\u00c8",
        "ÃŠ": "\u00ca",
        "Ã": "\u00cd",
        "ÃŒ": "\u00cc",
        "Ã“": "\u00d3",
        "Ã’": "\u00d2",
        "Ã”": "\u00d4",
        "Ã•": "\u00d5",
        "Ãš": "\u00da",
        "Ã™": "\u00d9",
        "Ã‡": "\u00c7",
        "Ã¡": "\u00e1",
        "Ã ": "\u00e0",
        "Ã¢": "\u00e2",
        "Ã£": "\u00e3",
        "Ã©": "\u00e9",
        "Ãª": "\u00ea",
        "Ã­": "\u00ed",
        "Ã³": "\u00f3",
        "Ãµ": "\u00f5",
        "Ãº": "\u00fa",
        "Ã§": "\u00e7",
    }

    for errado, certo in trocas.items():
        t = t.replace(errado, certo)

    return t


def _remover_fragmentos_soltos_saida(texto: str) -> str:
    t = str(texto)

    t = re.sub(r"\b(ATT|OBS|REF|CLI|CB|C\.B|PAQ|PAQ24|BULTO|CODIGO|C\u00d3DIGO)\b.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(TEL|TELEFONE|TELEMOVEL|TELEM\u00d3VEL|NIF|NIPC|CLIENTE)\b.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(3800|3810)\s*[- ]?\s*\d{3}\b.*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(3800|3810)\b\s*$", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(?:\+?351)?\s*9(?:\s*\d){8}\b.*$", "", t)
    t = re.sub(r"\b\d{6,}\b.*$", "", t)
    t = re.sub(
        r"\s+\b(?!LOTE|BLOCO|PORTA|PISO|ANDAR|ARMAZEM|ARMAZ\u00c9M)[A-Z]{2,4}\s+\d{3,}\b\s*$",
        "",
        t,
        flags=re.IGNORECASE,
    )

    return t.strip(" ,.;:-")


def _normalizar_numero_porta_morada(texto: str) -> str:
    t = str(texto)
    grau = "\u00ba"
    grau_alt = "\u00b0"
    grau_mojibake = "\u00c2\u00ba"
    simbolos = re.escape(grau + grau_alt + grau_mojibake)
    marcador = "N" + grau

    t = re.sub(rf"\bN\s*[{simbolos}]?\s*(\d{{1,5}}[A-Z]?)\b", rf"{marcador} \1", t, flags=re.IGNORECASE)
    t = re.sub(r"\bNR\s*(\d{1,5}[A-Z]?)\b", rf"{marcador} \1", t, flags=re.IGNORECASE)
    t = re.sub(rf"\bN\s*[{simbolos}]\s*(\d{{1,5}})(ESQ|DTO|DIR|DRT|FTE|FRENTE|TRAS|TRASEIRAS)\b", rf"{marcador} \1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\b(\d{1,5})(ESQ|DTO|DIR|DRT|FTE|FRENTE|TRAS|TRASEIRAS)\b", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\bR\s*/\s*C\b", "R/C", t, flags=re.IGNORECASE)

    return re.sub(r"\s+", " ", t).strip()


def _normalizar_mojibake_saida(texto: str) -> str:
    t = str(texto)
    trocas = {
        "N" + "\u00c2\u00ba": "N\u00ba",
        "N" + "\u00c2\u00b0": "N\u00ba",
        "\u00c3\u0081": "\u00c1",
        "\u00c3\u0080": "\u00c0",
        "\u00c3\u0082": "\u00c2",
        "\u00c3\u0083": "\u00c3",
        "\u00c3\u0089": "\u00c9",
        "\u00c3\u0088": "\u00c8",
        "\u00c3\u008a": "\u00ca",
        "\u00c3\u008d": "\u00cd",
        "\u00c3\u008c": "\u00cc",
        "\u00c3\u0093": "\u00d3",
        "\u00c3\u0092": "\u00d2",
        "\u00c3\u0094": "\u00d4",
        "\u00c3\u0095": "\u00d5",
        "\u00c3\u009a": "\u00da",
        "\u00c3\u0099": "\u00d9",
        "\u00c3\u0087": "\u00c7",
        "\u00c3\u00a1": "\u00e1",
        "\u00c3\u00a0": "\u00e0",
        "\u00c3\u00a2": "\u00e2",
        "\u00c3\u00a3": "\u00e3",
        "\u00c3\u00a9": "\u00e9",
        "\u00c3\u00aa": "\u00ea",
        "\u00c3\u00ad": "\u00ed",
        "\u00c3\u00b3": "\u00f3",
        "\u00c3\u00b5": "\u00f5",
        "\u00c3\u00ba": "\u00fa",
        "\u00c3\u00a7": "\u00e7",
    }

    for errado, certo in trocas.items():
        t = t.replace(errado, certo)

    return t


def _formatar_capitalizacao_morada(texto: str) -> str:
    artigos = {"DA", "DE", "DO", "DAS", "DOS", "E"}
    abreviaturas = {"Nº", "NÂº", "ESQ", "DTO", "DIR", "DRT", "CV", "RC", "R/C", "LT"}
    partes = []

    for token in str(texto).split():
        token_limpo = token.strip()
        upper = token_limpo.upper()

        if upper in artigos:
            partes.append(upper.lower())
        elif upper in abreviaturas or upper.startswith("Nº") or upper.startswith("NÂº"):
            partes.append(token_limpo.replace("NÂº", "Nº").replace("NÂ°", "Nº"))
        elif re.fullmatch(r"\d+(?:[A-Z]|\.[A-Z])?", upper):
            partes.append(upper)
        else:
            partes.append(token_limpo[:1].upper() + token_limpo[1:].lower())

    return " ".join(partes)


# =========================
# PALAVRAS
# =========================

PALAVRAS_MORADA = [
    "RUA", "R.", "AVENIDA", "AV.", "ALAMEDA", "TRAVESSA", "TV.",
    "LARGO", "PRAÇA", "PRACA", "PRACETA", "PCT", "ESTRADA",
    "ESTRADA NACIONAL", "EST.", "EST ", "ESTR.", "EST.NAC",
    "EST NAC", "NACIONAL", "CAMINHO", "CM.", "URBANIZAÇÃO",
    "URBANIZACAO", "ROTUNDA", "BECO", "BAIRRO", "QUINTA",
    "LOTE", "ZONA", "LUGAR", "CASAL", "VIVENDA", "CASA",
]

PALAVRAS_DESCARTAR = [
    "HTTP", "HTTPS", "APP.COM", "WWW", "ATT:", "OBS",
    "PROCURAR", "YVES", "ROCHER", "COSMETICOS", "COSMÉTICOS",
    "LDA", "S.A", " SA ", "000030038", "SN1", "SNI", "R-",
    "PALPITE", "EXP:", "REF:", "COD BULTO", "BULTO", "PESO",
    "DATA", "FECHA", "REMITENTE", "AMAZON", "SPAIN", "MADRID",
    "TIPO PORTES", "PAGADO", "REEMBOLSO", "ENVIO",
    "1DE1", "1 DE1", "KGS",
]

PALAVRAS_CORTE_ETIQUETA = [
    "C.B", "CB", "CLI", "UL.CLI", "OBS", "REF", "ATT", "PAQ",
    "PAQ24", "BULTO", "CODIGO", "CÓDIGO", "COD.", "COD",
    "CODIGO BULTO", "CÓDIGO BULTO", "EXP", "EXPEDICAO",
    "EXPEDIÇÃO", "REMETENTE", "DESTINATARIO", "DESTINATÁRIO",
]

PALAVRAS_FORMATO_MORADA = [
    "RUA", "AVENIDA", "ALAMEDA", "TRAVESSA", "LARGO", "PRAÇA",
    "PRACA", "PRACETA", "ESTRADA", "CAMINHO", "BECO", "BAIRRO",
    "QUINTA", "ROTUNDA", "LUGAR", "CASAL", "ZONA", "INDUSTRIAL",
    "URBANIZACAO", "URBANIZAÇÃO", "LOTE", "BLOCO", "PORTA",
    "PISO", "ANDAR", "ESQ", "DTO", "FRENTE", "ARMAZEM",
    "ARMAZÉM", "NACIONAL", "EUROPA", "REPUBLICA", "REPÚBLICA",
    "TABOEIRA",
]


# =========================
# CÓDIGO POSTAL
# =========================

def codigo_postal_valido(cp: str) -> bool:
    """
    Validação FINAL.
    Usada para confirmar e exportar.
    Aqui NÃO aceita apenas 3800/3810.
    """
    cp = str(cp).strip()

    if not re.match(r"^\d{4}-\d{3}$", cp):
        return False

    prefixo = cp[:4]
    sufixo = cp[5:]

    if prefixo not in PREFIXOS_ACEITES:
        return False

    if int(sufixo) < 1:
        return False

    return True


def codigo_postal_ocr_aceitavel(cp: str) -> bool:
    """
    Validação TEMPORÁRIA para leitura OCR.
    Aceita 3800/3810 apenas para mostrar resultado parcial,
    mas NÃO serve para confirmar/exportar.
    """
    cp = str(cp).strip()

    if re.match(r"^\d{4}$", cp):
        return cp in PREFIXOS_ACEITES

    return codigo_postal_valido(cp)


def codigo_postal_tem_sufixo(cp: str) -> bool:
    return bool(re.match(r"^\d{4}-\d{3}$", str(cp).strip()))


def texto_tem_localidade_aveiro(texto: str) -> bool:
    t = corrigir_ocr_para_morada(texto).upper()
    return any(loc in t for loc in LOCALIDADES_AVEIRO)


def linha_tem_lixo_para_cp(linha: str) -> bool:
    l = str(linha).upper()

    bloqueios = [
        "HTTP", "APP.COM", "WWW", "ATT:", "OBS", "PROCURAR",
        "PALPITE", "000030038", "EXP:", "REF:", "BULTO",
        "PESO", "COD.", "COD ", "GR:", "NºVEND", "VEND:",
        "DATA:", "HORA:",
        "C.B:", "CB:", "C.D", "CD:", "UL.CLI", "CLI",
        "PAQ", "PAQ24", "63-PAQ", "63PAC",
    ]

    return any(x in l for x in bloqueios)


# =========================
# MORADAS NÃO-PORTUGUESAS
# =========================

def eh_morada_nao_portugal(linha: str) -> bool:
    l = corrigir_ocr_para_morada(str(linha)).upper()

    if any(p in l for p in PALAVRAS_NAO_PORTUGAL):
        return True

    for cidade in CIDADES_NAO_PORTUGAL:
        if re.search(r"\b" + re.escape(cidade) + r"\b", l):
            return True

    if not re.search(r"\b\d{4}-\d{3}\b", l):
        m5 = re.search(r"\b(\d{5})\b", l)

        if m5:
            cp5 = m5.group(1)
            prefixo2 = cp5[:2]

            if prefixo2 in PREFIXOS_CP_ESPANHA:
                return True

    return False


# =========================
# GEOCODER
# =========================

def validar_morada_online(morada: str, codigo_postal: str) -> dict:
    resultado = {
        "valida": False,
        "servico": "nominatim",
        "display_name": "",
        "motivo": "",
    }

    if not GEOCODER_ENABLED:
        resultado["motivo"] = "geocoder_desativado"
        return resultado

    morada = formatar_morada_para_saida(morada)
    codigo_postal = str(codigo_postal).strip()

    if not morada or morada == "Não encontrada":
        resultado["motivo"] = "morada_vazia"
        return resultado

    if not codigo_postal_ocr_aceitavel(codigo_postal):
        resultado["motivo"] = "codigo_postal_invalido"
        return resultado

    if eh_morada_nao_portugal(morada):
        resultado["motivo"] = "morada_nao_portugal"
        return resultado

    try:
        query = f"{morada}, {codigo_postal}, Aveiro, Portugal"

        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": query,
                "format": "json",
                "addressdetails": 1,
                "limit": 5,
                "countrycodes": "pt",
            },
            headers={
                "User-Agent": GEOCODER_USER_AGENT,
            },
            timeout=GEOCODER_TIMEOUT,
        )

        if response.status_code != 200:
            resultado["motivo"] = f"status_{response.status_code}"
            return resultado

        dados = response.json()

        if not dados:
            resultado["motivo"] = "sem_resultados"
            return resultado

        prefixo_cp = codigo_postal[:4]

        for item in dados:
            display = str(item.get("display_name", "")).upper()
            address = item.get("address", {}) or {}

            postcode = str(address.get("postcode", "")).strip()
            city = str(
                address.get("city")
                or address.get("town")
                or address.get("village")
                or address.get("municipality")
                or address.get("county")
                or ""
            ).upper()

            country_code = str(address.get("country_code", "")).lower()

            bate_pais = country_code == "pt" or "PORTUGAL" in display
            bate_aveiro = "AVEIRO" in display or "AVEIRO" in city
            bate_cp = (
                postcode.startswith(prefixo_cp)
                or prefixo_cp in display
                or postcode.startswith("3800")
                or postcode.startswith("3810")
            )

            if bate_pais and bate_aveiro and bate_cp:
                resultado["valida"] = True
                resultado["display_name"] = item.get("display_name", "")
                resultado["motivo"] = "validada"
                return resultado

        resultado["motivo"] = "resultado_nao_bate_aveiro_cp"
        resultado["display_name"] = dados[0].get("display_name", "")

        return resultado

    except requests.exceptions.Timeout:
        resultado["motivo"] = "timeout_geocoder"
        return resultado

    except Exception as e:
        resultado["motivo"] = f"erro_geocoder: {str(e)}"
        return resultado


# =========================
# MORADA
# =========================

def eh_linha_cidade(linha: str) -> bool:
    l = corrigir_ocr_para_morada(linha).upper().strip()

    if not l:
        return False

    if any(x in l for x in PALAVRAS_DESCARTAR):
        return False

    if re.search(r"\d{4}-\d{3}", l):
        return False

    letras = len(re.findall(r"[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]", l))
    numeros = len(re.findall(r"\d", l))

    return letras >= 3 and numeros <= 3


def linha_tem_via(linha: str) -> bool:
    l = normalizar_morada_extraida(linha).upper()
    return any(p in l for p in PALAVRAS_MORADA)


def eh_morada_valida(linha: str) -> bool:
    l = normalizar_morada_extraida(linha).upper().strip()

    if not l:
        return False

    if any(x in l for x in PALAVRAS_DESCARTAR):
        return False

    if re.search(r"\d{4}-\d{3}", l):
        return False

    tem_palavra_morada = any(p in l for p in PALAVRAS_MORADA)
    tem_numero = bool(re.search(r"\d", l))

    if tem_palavra_morada and tem_numero:
        return True

    if tem_palavra_morada and ("Nº" in l or " N " in l or " NR " in l):
        return True

    if tem_palavra_morada and len(l) >= 12:
        return True

    complementos = [
        "VIVENDA", "PORTÃO", "PORTAO", "VERMELHO", "GRANDE",
        "LOTE", "BLOCO", "ANDAR", "ESQ", "DTO", "FRENTE",
        "CASA", "ARMAZEM", "ARMAZÉM", "PORTA", "PISO",
    ]

    if any(c in l for c in complementos) and len(l) >= 8:
        return True

    return False


def limpar_morada_final(morada: str) -> str:
    m = normalizar_morada_extraida(morada)

    vias = [
        "RUA", "AVENIDA", "ALAMEDA", "TRAVESSA", "LARGO",
        "PRAÇA", "PRACA", "PRACETA", "ESTRADA", "CAMINHO",
        "BECO", "BAIRRO", "QUINTA", "ROTUNDA", "LUGAR",
        "CASAL", "ZONA",
    ]

    upper = m.upper()
    melhor_pos = None

    for via in vias:
        pos = upper.find(via)

        if pos != -1:
            if melhor_pos is None or pos < melhor_pos:
                melhor_pos = pos

    if melhor_pos is not None:
        m = m[melhor_pos:].strip()

    # Corrige N418 / N 418 / Nº418
    m = re.sub(r"\bN\s*º?\s*(\d{1,5})\b", r"Nº \1", m, flags=re.IGNORECASE)

    # Remove telefone/tracking depois do número da porta
    # Ex: AVENIDA EUROPA Nº 418 965750109 -> AVENIDA EUROPA Nº 418
    m = re.sub(
        r"\b(Nº\s*\d{1,5}[A-Z]?)\s+\d{6,}\b.*$",
        r"\1",
        m,
        flags=re.IGNORECASE,
    )

    # Remove qualquer número gigante que ficou agarrado à morada
    m = re.sub(r"\b\d{6,}\b.*$", "", m).strip()

    # Remove COD. AT / telefone / referências depois da morada
    m = re.sub(
        r"\b(COD\.?\s*AT|CODIGO|CÓDIGO|TEL|TELEMOVEL|TELEMÓVEL|NIF|NIPC|CLIENTE|CLI)\b.*$",
        "",
        m,
        flags=re.IGNORECASE,
    ).strip()

    # Remove CP completo e tudo depois dele
    m = re.sub(r"\b(3800|3810)\s*[- ]?\s*\d{3}\b.*$", "", m).strip()

    # Remove CP incompleto quando ficou no fim da morada
    m = re.sub(r"\b(3800|3810)\b\s*$", "", m).strip()

    lixo = [
        "DESTINATARIO", "DESTINATÁRIO", "REMETENTE", "REMITENTE",
        "OBSERVACIONES", "OBSERVAÇÕES", "OBSERVACOES",
        "VALORES AÑADIDOS", "VALORES ADICIONADOS",
        "CODIGO BULTO", "CÓDIGO BULTO",
        "ATT", "OBS",
    ]

    for palavra in lixo:
        m = re.sub(r"\b" + re.escape(palavra) + r"\b", "", m, flags=re.IGNORECASE)

    m = re.sub(r"\s+", " ", m).strip()
    m = m.strip(" ,.;:-")

    return m


def formatar_morada_para_saida(morada: str) -> str:
    """
    Limpeza/formatação final genérica para mostrar, confirmar e exportar.
    A função não cria moradas novas; só trabalha com o texto encontrado/editado.
    """
    if morada is None:
        return ""

    original = str(morada).strip()

    if not original:
        return ""

    if original.upper() in {"NÃO ENCONTRADA", "NAO ENCONTRADA", "NÃO ENCONTRADO", "NAO ENCONTRADO"}:
        return "Não encontrada"

    m = original.replace("\r", " ").replace("\n", " ")
    m = re.sub(r"[|_]+", " ", m)
    m = re.sub(r"\s+", " ", m).strip()

    m = _remover_lixo_etiqueta_morada(m)
    m = normalizar_morada_extraida(m)
    m = _corrigir_tokens_ocr_morada(m)
    m = _separar_palavras_coladas_morada(m)
    m = _normalizar_numero_porta_morada(m)
    m = _remover_lixo_etiqueta_morada(m)
    m = limpar_morada_final(m)
    m = _normalizar_numero_porta_morada(m)
    m = _remover_fragmentos_soltos_saida(m)

    m = re.sub(r"\bN\s*[º°]?\s*(\d{1,5}[A-Z]?)\b", r"Nº \1", m, flags=re.IGNORECASE)
    m = re.sub(r"\bNR\s*(\d{1,5}[A-Z]?)\b", r"Nº \1", m, flags=re.IGNORECASE)
    m = re.sub(r"\bN\s+º\s*", "Nº ", m, flags=re.IGNORECASE)
    m = re.sub(r"\bN[º°]{1,3}\s*", "Nº ", m, flags=re.IGNORECASE)
    m = re.sub(r"\s*([,.;:])\s*", r"\1 ", m)
    m = re.sub(r"\b(\d{1,4})\.\s+([A-Z])\b", r"\1.\2", m, flags=re.IGNORECASE)
    m = re.sub(r"\s+([A-Z])\s*/\s*([A-Z])\b", r" \1/\2", m, flags=re.IGNORECASE)
    m = re.sub(r"\s+", " ", m).strip(" ,.;:-")

    if not m:
        return ""

    return _normalizar_mojibake_saida(_formatar_capitalizacao_morada(m))


def linha_parece_complemento_morada(linha: str) -> bool:
    l = normalizar_morada_extraida(linha).upper()

    if not l:
        return False

    if any(x in l for x in PALAVRAS_DESCARTAR):
        return False

    if _linha_quebra_bloco(linha):
        return False

    complementos = [
        "VIVENDA", "PORTÃO", "PORTAO", "VERMELHO", "GRANDE",
        "LOTE", "BLOCO", "ANDAR", "ESQ", "DTO", "FRENTE",
        "CASA", "ARMAZEM", "ARMAZÉM", "PORTA", "PISO",
        "COM ", "JUNTO", "TRASEIRAS",
    ]

    if any(c in l for c in complementos):
        return True

    letras = len(re.findall(r"[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]", l))
    numeros = len(re.findall(r"\d", l))

    if letras >= 8 and numeros <= 4 and not linha_tem_via(l):
        return True

    return False


def linha_ruim_para_morada(linha: str) -> bool:
    l = str(linha).upper().strip()

    if not l:
        return True

    bloqueios = [
        "C.B:", "CB:", "CODIGO BULTO", "CÓDIGO BULTO",
        "OBS", "EXPEDICION", "EXPEDIÇÃO", "REMETENTE",
        "REMITENTE", "PAQ", "PAQ24", "PESO", "KGS", "REF:",
        "HTTP", "WWW", "VALORES", "BARCELONA", "MADRID",
        "SPAIN", "ESPAÑA", "ESPANA", "COD. AT", "COD AT",
        "BUL.", "CLI", "UL.CLI",
    ]

    if any(b in l for b in bloqueios):
        return True

    if re.match(r"^[A-Z0-9]{12,}$", l):
        return True

    if re.match(r"^\d{7,}$", l):
        return True

    numeros = len(re.findall(r"\d", l))
    letras = len(re.findall(r"[A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ]", l))

    if numeros >= 8 and letras <= 3:
        return True

    return False


def extrair_morada_antes_do_cp_na_linha(linha: str, inicio_cp: int) -> str:
    parte_antes = str(linha)[:inicio_cp].strip()

    if not parte_antes:
        return ""

    parte_antes = limpar_morada_final(parte_antes)

    if eh_morada_nao_portugal(parte_antes):
        return ""

    if eh_morada_valida(parte_antes):
        return parte_antes

    if linha_tem_via(parte_antes) and re.search(r"\d", parte_antes):
        return parte_antes

    return ""


def pontuar_morada(linha: str, index: int, cp_index: int) -> int:
    l = normalizar_morada_extraida(linha).upper()
    score = 0

    if eh_morada_nao_portugal(linha):
        return -9999

    if eh_morada_valida(linha):
        score += 150

    if "AVENIDA" in l or "AV." in l:
        score += 35

    if "RUA" in l or " R." in l:
        score += 35

    if "ALAMEDA" in l:
        score += 25

    if "TRAVESSA" in l:
        score += 20

    if "PRACETA" in l or "PCT" in l:
        score += 20

    if "ESTRADA" in l or "NACIONAL" in l or "EST.NAC" in l:
        score += 30

    if "Nº" in l or "N°" in l:
        score += 18

    if re.search(r"\d", l):
        score += 20

    distancia = abs(cp_index - index)
    score += max(0, 70 - distancia * 10)

    if index < cp_index:
        score += 30

    if index > cp_index:
        score -= 20

    if index < cp_index - 8:
        score -= 35

    if texto_tem_localidade_aveiro(l):
        score += 40

    return score


def _linha_quebra_bloco(linha: str) -> bool:
    l = linha.strip().upper()

    if not l:
        return True

    if re.match(r"^\d{8,}$", l):
        return True

    if re.match(r"^[A-Z]{2}\d{9,}[A-Z]{2}$", l):
        return True

    if re.match(r"^[A-Z0-9]{14,}$", l):
        return True

    LOGISTICA = [
        "REMITENTE", "DESTINATARIO", "DESTINATÁRIO",
        "TIPO PORTES", "REEMBOLSO", "PAGADO", "BULTO",
        "COD BULTO", "1DE1", "1 DE 1", "1 DE1",
        "PESO:", "KGS", "ENVIO", "FECHA",
        "ATT:", "EXP:", "REF:", "OBS:",
        "C.B:", "CB:", "UL.CLI",
    ]

    if any(k in l for k in LOGISTICA):
        return True

    if re.search(r"\b\d{4}[- ]\d{3}\b", l):
        return True

    if re.search(r"(?<!\d)\d{5}(?!\d)", l) and not re.search(r"\b\d{4}[- ]\d{3}\b", l):
        return True

    return False


def montar_morada_multilinha(linhas: list[str], cp_index: int) -> str:
    candidatos = []
    inicio = max(0, cp_index - 8)

    for i in range(cp_index - 1, inicio - 1, -1):
        linha_base_raw = linhas[i].strip()

        if linha_ruim_para_morada(linha_base_raw):
            continue

        linha_base = normalizar_morada_extraida(linha_base_raw)

        if eh_morada_nao_portugal(linha_base):
            continue

        if not linha_tem_via(linha_base) and not eh_morada_valida(linha_base):
            continue

        partes = [linha_base]

        for j in range(i + 1, min(cp_index, i + 4)):
            prox_raw = linhas[j].strip()

            if linha_ruim_para_morada(prox_raw):
                continue

            if _linha_quebra_bloco(prox_raw):
                continue

            prox = normalizar_morada_extraida(prox_raw)

            if eh_morada_nao_portugal(prox):
                continue

            if linha_parece_complemento_morada(prox) or linha_tem_via(prox):
                partes.append(prox)

        combinada = " ".join(partes)
        combinada = limpar_morada_final(combinada)

        if not combinada:
            continue

        if not eh_morada_valida(combinada):
            continue

        score = 200
        score += max(0, 80 - abs(cp_index - i) * 10)

        if "Nº" in combinada:
            score += 40

        if re.search(r"\d", combinada):
            score += 35

        if any(c in combinada.upper() for c in ["VIVENDA", "PORTÃO", "VERMELHO", "GRANDE", "LOTE", "ESQ", "DTO"]):
            score += 30

        candidatos.append({
            "morada": combinada,
            "score": score,
        })

    if not candidatos:
        return ""

    candidatos.sort(key=lambda x: x["score"], reverse=True)

    return candidatos[0]["morada"]


def formatar_morada_pelo_geocoder(morada: str, codigo_postal: str, geo: dict) -> str:
    if not geo.get("valida"):
        return morada

    display = geo.get("display_name", "") or ""

    if not display:
        return morada

    primeira_parte = display.split(",")[0].strip()

    if not primeira_parte:
        return morada

    morada_limpa = formatar_morada_para_saida(morada)
    primeira_parte = normalizar_morada_extraida(primeira_parte)

    if linha_tem_via(primeira_parte) and len(primeira_parte) >= 8:
        numeros = re.findall(r"\b\d+[A-Z]?\b", morada_limpa)
        complemento = ""

        for palavra in ["VIVENDA", "PORTÃO", "VERMELHO", "GRANDE", "LOTE", "BLOCO", "ESQ", "DTO"]:
            if palavra in morada_limpa.upper():
                complemento = morada_limpa
                break

        if numeros and not re.search(r"\b\d+[A-Z]?\b", primeira_parte):
            primeira_parte = f"{primeira_parte} Nº {numeros[0]}"

        if complemento and complemento.upper() != primeira_parte.upper():
            return formatar_morada_para_saida(morada_limpa)

        return formatar_morada_para_saida(primeira_parte)

    return formatar_morada_para_saida(morada_limpa)


# =========================
# EXTRAÇÃO DE CÓDIGO POSTAL
# =========================

def extrair_cp_mesma_linha(linha: str) -> dict | None:
    original = str(linha).upper()

    if linha_tem_lixo_para_cp(original):
        return None

    linha_corrigida = corrigir_ocr_para_morada(original)
    linha_num = converter_ocr_numero(linha_corrigida)
    linha_num = linha_num.replace("PT-", "").replace("P-", "").replace("P ", "")

    m = re.search(r"(?<!\d)(3800|3810)\s*[- ]\s*(\d{3})(?!\d)", linha_num)

    if m:
        cp = f"{m.group(1)}-{m.group(2)}"

        if codigo_postal_valido(cp):
            cidade = linha_num[m.end():].strip()
            cidade = corrigir_ocr_para_morada(cidade)
            cidade = re.sub(r"[^A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]", " ", cidade)
            cidade = limpar_linha(cidade)

            morada_na_linha = extrair_morada_antes_do_cp_na_linha(
                linha_corrigida,
                m.start(),
            )

            return {
                "codigo": cp,
                "linha_index": None,
                "cidade": cidade,
                "morada_na_linha": morada_na_linha,
                "origem": "mesma_linha",
            }

    m_compact = re.search(r"(?<!\d)(3800|3810)(\d{3})(?!\d)", linha_num)

    if m_compact:
        cp = f"{m_compact.group(1)}-{m_compact.group(2)}"

        if codigo_postal_valido(cp):
            cidade = linha_num[m_compact.end():].strip()
            cidade = corrigir_ocr_para_morada(cidade)
            cidade = re.sub(r"[^A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]", " ", cidade)
            cidade = limpar_linha(cidade)

            morada_na_linha = extrair_morada_antes_do_cp_na_linha(
                linha_corrigida,
                m_compact.start(),
            )

            return {
                "codigo": cp,
                "linha_index": None,
                "cidade": cidade,
                "morada_na_linha": morada_na_linha,
                "origem": "compacto",
            }

    m_3801 = re.search(r"(?<!\d)3801\s*[- ]\s*(\d{3})(?!\d)", linha_num)

    if m_3801 and any(x in original for x in ["AVEIRO", "AVRO", "AVR0", "AVEI", "ESGUEIRA"]):
        cp = f"3810-{m_3801.group(1)}"

        if codigo_postal_valido(cp):
            cidade = linha_num[m_3801.end():].strip()
            cidade = corrigir_ocr_para_morada(cidade)
            cidade = re.sub(r"[^A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]", " ", cidade)
            cidade = limpar_linha(cidade)

            morada_na_linha = extrair_morada_antes_do_cp_na_linha(
                linha_corrigida,
                m_3801.start(),
            )

            return {
                "codigo": cp,
                "linha_index": None,
                "cidade": cidade,
                "morada_na_linha": morada_na_linha,
                "origem": "corrigido_3801",
            }

    return None


def extrair_cp_partido(linhas: list[str], i: int) -> dict | None:
    linha = linhas[i].strip()

    if linha_tem_lixo_para_cp(linha):
        return None

    linha_num = converter_ocr_numero(linha)
    linha_limpa = re.sub(r"[^0-9]", "", linha_num)

    if linha_limpa not in ["3800", "3810", "3801"]:
        return None

    prefixo = "3810" if linha_limpa == "3801" else linha_limpa

    if i + 1 >= len(linhas):
        return {
            "codigo": prefixo,
            "linha_index": i,
            "cidade": "",
            "morada_na_linha": "",
            "origem": "prefixo_sem_sufixo",
        }

    prox_original = linhas[i + 1].strip()

    if linha_tem_lixo_para_cp(prox_original):
        return {
            "codigo": prefixo,
            "linha_index": i,
            "cidade": "",
            "morada_na_linha": "",
            "origem": "prefixo_sem_sufixo",
        }

    prox_corrigida = corrigir_ocr_para_morada(prox_original)
    prox_num = converter_ocr_numero(prox_corrigida)

    m = re.match(r"^\s*(\d{3})\s*[- ]?\s*(.*)$", prox_num)

    if m:
        cp = f"{prefixo}-{m.group(1)}"

        if codigo_postal_valido(cp):
            cidade = m.group(2).strip()
            cidade = corrigir_ocr_para_morada(cidade)
            cidade = re.sub(r"[^A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]", " ", cidade)
            cidade = limpar_linha(cidade)

            return {
                "codigo": cp,
                "linha_index": i,
                "cidade": cidade,
                "morada_na_linha": "",
                "origem": "partido",
            }

    if texto_tem_localidade_aveiro(prox_corrigida):
        m_num = re.search(r"(?<!\d)(\d{3})(?!\d)", prox_num)

        if m_num:
            suffix = m_num.group(1)
            cidade = prox_corrigida.replace(suffix, "")
            cidade = re.sub(r"[^A-ZÁÀÂÃÉÈÊÍÌÓÒÔÕÚÙÇ ]", " ", cidade.upper())
            cidade = limpar_linha(cidade)

            cp = f"{prefixo}-{suffix}"

            if codigo_postal_valido(cp):
                return {
                    "codigo": cp,
                    "linha_index": i,
                    "cidade": cidade,
                    "morada_na_linha": "",
                    "origem": "partido_localidade",
                }

        return {
            "codigo": prefixo,
            "linha_index": i,
            "cidade": limpar_linha(corrigir_ocr_para_morada(prox_corrigida)),
            "morada_na_linha": "",
            "origem": "prefixo_com_localidade",
        }

    return {
        "codigo": prefixo,
        "linha_index": i,
        "cidade": "",
        "morada_na_linha": "",
        "origem": "prefixo_sem_sufixo",
    }


def extrair_codigos_postais_aveiro(linhas: list[str]) -> list[dict]:
    encontrados = []

    for i, linha in enumerate(linhas):
        mesmo = extrair_cp_mesma_linha(linha)

        if mesmo:
            cidade = mesmo.get("cidade", "")

            if not cidade and i + 1 < len(linhas):
                prox = linhas[i + 1].strip()

                if eh_linha_cidade(prox):
                    cidade = prox

            encontrados.append({
                "codigo": mesmo["codigo"],
                "linha_index": i,
                "cidade": limpar_linha(corrigir_ocr_para_morada(cidade)),
                "morada_na_linha": mesmo.get("morada_na_linha", ""),
                "origem": mesmo.get("origem", "mesma_linha"),
            })

        partido = extrair_cp_partido(linhas, i)

        if partido:
            encontrados.append(partido)

    melhores = {}

    for item in encontrados:
        chave = item["codigo"]

        if chave not in melhores:
            melhores[chave] = item
            continue

        atual = melhores[chave]

        atual_tem_morada = bool(atual.get("morada_na_linha"))
        novo_tem_morada = bool(item.get("morada_na_linha"))

        if novo_tem_morada and not atual_tem_morada:
            melhores[chave] = item

        if codigo_postal_valido(item["codigo"]) and not codigo_postal_valido(atual["codigo"]):
            melhores[chave] = item

    return list(melhores.values())


def encontrar_morada_para_codigo(linhas: list[str], cp_info: dict) -> str:
    morada_na_linha = cp_info.get("morada_na_linha", "")

    if morada_na_linha:
        morada_na_linha = limpar_morada_final(morada_na_linha)

        if morada_na_linha and not eh_morada_nao_portugal(morada_na_linha):
            return morada_na_linha

    cp_index = cp_info["linha_index"]

    morada_bloco = montar_morada_multilinha(linhas, cp_index)

    if morada_bloco and not eh_morada_nao_portugal(morada_bloco):
        return morada_bloco

    candidatos = []

    for i in range(cp_index - 1, max(-1, cp_index - 12), -1):
        linha_raw = linhas[i].strip()
        linha = normalizar_morada_extraida(linha_raw)

        if linha_ruim_para_morada(linha_raw):
            continue

        if eh_morada_nao_portugal(linha):
            continue

        if eh_morada_valida(linha):
            score = pontuar_morada(linha, i, cp_index)

            if score > 0:
                candidatos.append({
                    "linha": limpar_morada_final(linha),
                    "score": score,
                    "index": i,
                })

    if not candidatos:
        return ""

    candidatos.sort(key=lambda x: x["score"], reverse=True)

    return candidatos[0]["linha"]


def pontuar_resultado(resultado: dict) -> int:
    score = 0

    morada = resultado.get("morada", "")
    codigo = resultado.get("codigo_postal", "")
    cidade = resultado.get("cidade", "")
    contexto = resultado.get("contexto", "")
    origem_codigo = resultado.get("origem_codigo", "")

    if codigo_postal_valido(codigo):
        score += 260
    elif codigo_postal_ocr_aceitavel(codigo):
        score += 120

    if morada and morada != "Não encontrada":
        score += 220

    if eh_morada_valida(morada):
        score += 150

    if eh_morada_nao_portugal(morada):
        score -= 9999

    if origem_codigo in ["mesma_linha", "compacto", "corrigido_3801", "compacto_3801"]:
        if morada and morada != "Não encontrada":
            score += 300

    if origem_codigo in ["partido", "partido_localidade"]:
        score -= 40

    if origem_codigo in ["prefixo_sem_sufixo", "prefixo_com_localidade"]:
        score -= 80

    if cidade and cidade != "Não encontrada":
        score += 30

    if texto_tem_localidade_aveiro(cidade):
        score += 45

    if texto_tem_localidade_aveiro(contexto):
        score += 45

    texto_contexto = contexto.upper()

    if any(x in texto_contexto for x in ["R-", "SN1", "PALPITE", "ATT:", "OBS", "EXP:", "REF:", "BULTO", "C.B:", "CB:"]):
        score -= 70

    return score


def escolher_destinatario(resultados: list[dict]) -> dict | None:
    validos = [
        r for r in resultados
        if codigo_postal_ocr_aceitavel(r["codigo_postal"])
        and r.get("morada", "Não encontrada") != "Não encontrada"
        and not eh_morada_nao_portugal(r.get("morada", ""))
    ]

    if not validos:
        validos = [
            r for r in resultados
            if codigo_postal_ocr_aceitavel(r["codigo_postal"])
        ]

    if not validos:
        return None

    validos.sort(key=lambda x: x["score"], reverse=True)

    return validos[0]


def extrair_dados_aveiro(texto: str) -> dict:
    texto = normalizar_texto(texto)
    linhas = [l.strip() for l in texto.split("\n") if l.strip()]
    cps = extrair_codigos_postais_aveiro(linhas)
    resultados = []

    for cp in cps:
        if not codigo_postal_ocr_aceitavel(cp["codigo"]):
            continue

        morada = encontrar_morada_para_codigo(linhas, cp)
        morada = formatar_morada_para_saida(morada) if morada else ""
        cidade = corrigir_ocr_para_morada(cp.get("cidade", ""))
        idx = cp["linha_index"]

        contexto_inicio = max(0, idx - 6)
        contexto_fim = min(len(linhas), idx + 6)
        contexto = "\n".join(linhas[contexto_inicio:contexto_fim])

        item = {
            "morada": morada if morada else "Não encontrada",
            "codigo_postal": cp["codigo"],
            "cidade": cidade if cidade else "",
            "linha_codigo_index": idx,
            "origem_codigo": cp.get("origem", ""),
            "contexto": contexto,
            "geo_validada": False,
            "geo_motivo": "",
            "geo_display_name": "",
            "codigo_incompleto": not codigo_postal_valido(cp["codigo"]),
        }

        item["score"] = pontuar_resultado(item)
        resultados.append(item)

    resultados.sort(key=lambda x: x["score"], reverse=True)

    candidatos_geo = [
        r for r in resultados
        if r.get("morada") != "Não encontrada"
        and codigo_postal_ocr_aceitavel(r.get("codigo_postal", ""))
    ][:3]

    for item in candidatos_geo:
        geo = validar_morada_online(item["morada"], item["codigo_postal"])

        item["geo_validada"] = geo.get("valida", False)
        item["geo_motivo"] = geo.get("motivo", "")
        item["geo_display_name"] = geo.get("display_name", "")

        if geo.get("valida"):
            item["score"] += 250

            morada_formatada = formatar_morada_pelo_geocoder(
                item["morada"],
                item["codigo_postal"],
                geo,
            )

            if morada_formatada:
                item["morada"] = morada_formatada
        else:
            if GEOCODER_STRICT:
                item["score"] -= 180
            else:
                item["score"] -= 10

    resultados.sort(key=lambda x: x["score"], reverse=True)

    escolhido = escolher_destinatario(resultados)

    if escolhido:
        for item in resultados:
            if item.get("morada") and item.get("morada") != "Não encontrada":
                item["morada"] = formatar_morada_para_saida(item["morada"])

        return {
            "morada": formatar_morada_para_saida(escolhido["morada"]),
            "codigo_postal": escolhido["codigo_postal"],
            "cidade": escolhido.get("cidade", ""),
            "geo_validada": escolhido.get("geo_validada", False),
            "geo_motivo": escolhido.get("geo_motivo", ""),
            "geo_display_name": escolhido.get("geo_display_name", ""),
            "codigo_incompleto": not codigo_postal_valido(escolhido["codigo_postal"]),
            "todos_resultados": resultados,
        }

    return {
        "morada": "Não encontrada",
        "codigo_postal": "Não encontrado",
        "cidade": "",
        "geo_validada": False,
        "geo_motivo": "",
        "geo_display_name": "",
        "codigo_incompleto": False,
        "todos_resultados": resultados,
    }


# =========================
# IMAGEM
# =========================

def pre_processar_imagem(img: np.ndarray) -> np.ndarray:
    h, w = img.shape[:2]
    max_dim = 1600

    if max(h, w) > max_dim:
        scale = max_dim / max(h, w)
        img = cv2.resize(
            img,
            (int(w * scale), int(h * scale)),
            interpolation=cv2.INTER_AREA,
        )

    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)

    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
    l = clahe.apply(l)

    lab = cv2.merge((l, a, b))
    img = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)

    return img


def criar_versoes_imagem(caminho: str) -> list[str]:
    img = cv2.imread(caminho)

    if img is None:
        return [caminho]

    base = str(uuid.uuid4())
    versoes = []

    img_proc = pre_processar_imagem(img.copy())
    proc_path = f"uploads/proc_{base}.jpg"

    cv2.imwrite(proc_path, img_proc, [cv2.IMWRITE_JPEG_QUALITY, 95])
    versoes.append(proc_path)

    gray = cv2.cvtColor(img_proc, cv2.COLOR_BGR2GRAY)

    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)

    kernel = np.array([
        [0, -1, 0],
        [-1, 5, -1],
        [0, -1, 0],
    ])

    gray = cv2.filter2D(gray, -1, kernel)

    gray_path = f"uploads/gray_{base}.jpg"
    cv2.imwrite(gray_path, gray, [cv2.IMWRITE_JPEG_QUALITY, 95])
    versoes.append(gray_path)

    blur = cv2.GaussianBlur(gray, (3, 3), 0)

    binary = cv2.adaptiveThreshold(
        blur,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        25,
        8,
    )

    bin_path = f"uploads/bin_{base}.jpg"
    cv2.imwrite(bin_path, binary, [cv2.IMWRITE_JPEG_QUALITY, 95])
    versoes.append(bin_path)

    return versoes


# =========================
# OCR PARSER
# =========================

def extrair_texto_resultado_ocr(resultado) -> str:
    textos = []

    if not resultado:
        return ""

    if isinstance(resultado, list):
        for item in resultado:
            if isinstance(item, dict):
                if "rec_texts" in item and isinstance(item["rec_texts"], list):
                    textos.extend([str(x) for x in item["rec_texts"]])
                elif "text" in item:
                    textos.append(str(item["text"]))

            elif isinstance(item, list):
                for linha in item:
                    try:
                        if isinstance(linha, (list, tuple)) and len(linha) >= 2:
                            data = linha[1]

                            if isinstance(data, (tuple, list)):
                                textos.append(str(data[0]))
                            else:
                                textos.append(str(data))
                    except Exception:
                        pass

    return "\n".join(textos)


def juntar_textos_unicos(textos: list[str]) -> str:
    linhas_final = []
    vistos = set()

    for texto in textos:
        texto = normalizar_texto(texto)

        for linha in texto.split("\n"):
            linha = limpar_linha(linha)

            if not linha:
                continue

            chave = linha.upper()

            if chave not in vistos:
                vistos.add(chave)
                linhas_final.append(linha)

    return "\n".join(linhas_final)


def _ocr_uma_versao(path: str) -> str:
    engine = get_ocr()

    try:
        resultado = engine.ocr(path, cls=True)
        texto = extrair_texto_resultado_ocr(resultado)

        return normalizar_texto(texto)

    except TypeError:
        try:
            resultado = engine.ocr(path)
            texto = extrair_texto_resultado_ocr(resultado)

            return normalizar_texto(texto)

        except Exception:
            traceback.print_exc()
            return ""

    except Exception:
        traceback.print_exc()
        return ""


def rodar_ocr_em_versoes(versoes: list[str]) -> str:
    textos = []

    for i, path in enumerate(versoes):
        print(f"OCR versão {i + 1}/{len(versoes)}: {path}", flush=True)

        texto = _ocr_uma_versao(path)

        if texto:
            print(f"Texto versão {i + 1}:\n{texto}", flush=True)
            textos.append(texto)

    return juntar_textos_unicos(textos)


# =========================
# EXPORTAÇÃO
# =========================

def salvar_lote_em_arquivos(snapshot=None, excel_path=EXPORT_EXCEL, csv_path=EXPORT_CSV):
    origem = snapshot if snapshot is not None else list(lote_confirmado)
    dados = []

    for item in origem:
        item_export = dict(item)
        item_export["Morada"] = formatar_morada_para_saida(item_export.get("Morada", ""))
        item_export["Código Postal"] = str(item_export.get("Código Postal", "")).strip()

        if not codigo_postal_valido(item_export["Código Postal"]):
            raise ValueError(
                "Código postal incompleto ou inválido. Exporte apenas códigos no formato 3800-XXX ou 3810-XXX."
            )

        dados.append(item_export)

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Etiquetas Aveiro"

    COR_TITULO_BG = "0F172A"
    COR_HEADER_BG = "1D4ED8"
    COR_HEADER_FONT = "FFFFFF"
    COR_LINHA_PAR = "EEF6FF"
    COR_LINHA_IMPAR = "FFFFFF"
    COR_BORDA = "CBD5E1"
    COR_TEXTO = "0F172A"
    COR_MUTED = "64748B"

    borda = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    ws.merge_cells("A1:B1")
    titulo = ws["A1"]
    titulo.value = "Exportação de Etiquetas — Aveiro"
    titulo.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:B2")
    resumo = ws["A2"]
    resumo.value = f"Total de etiquetas exportadas: {len(dados)}"
    resumo.font = Font(name="Arial", size=11, color=COR_MUTED)
    resumo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    headers = ["Morada", "Código Postal"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, size=12, color=COR_HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    ws.row_dimensions[3].height = 28

    for idx, item in enumerate(dados, start=4):
        cor_bg = COR_LINHA_PAR if idx % 2 == 0 else COR_LINHA_IMPAR
        fill = PatternFill("solid", fgColor=cor_bg)

        morada = item.get("Morada", "")
        cp = item.get("Código Postal", "")

        c1 = ws.cell(row=idx, column=1, value=morada)
        c1.font = Font(name="Arial", size=11, color=COR_TEXTO)
        c1.fill = fill
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c1.border = borda

        c2 = ws.cell(row=idx, column=2, value=cp)
        c2.font = Font(name="Arial", size=11, bold=True, color="1D4ED8")
        c2.fill = fill
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = borda

        ws.row_dimensions[idx].height = 24

    ws.column_dimensions["A"].width = 75
    ws.column_dimensions["B"].width = 20

    ws.freeze_panes = "A4"

    if dados:
        tabela_ref = f"A3:B{len(dados) + 3}"
        tabela = Table(displayName="TabelaEtiquetasAveiro", ref=tabela_ref)

        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

    footer_row = len(dados) + 5

    ws.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=2,
    )

    footer = ws.cell(row=footer_row, column=1)
    footer.value = "Gerado automaticamente pelo sistema OCR de etiquetas"
    footer.font = Font(name="Arial", italic=True, size=10, color=COR_MUTED)
    footer.alignment = Alignment(horizontal="center")

    wb.save(excel_path)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Morada", "Código Postal"])

        for item in dados:
            writer.writerow([
                item.get("Morada", ""),
                item.get("Código Postal", ""),
            ])


def limpar_ficheiros_exportacao(*paths):
    for path in paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except Exception:
            pass


def limpar_pasta_exports():
    try:
        os.makedirs("exports", exist_ok=True)

        for nome in os.listdir("exports"):
            caminho = os.path.join("exports", nome)

            if os.path.isfile(caminho):
                os.remove(caminho)

    except Exception:
        pass


# =========================
# UPLOAD
# =========================

@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    caminhos_temporarios = []

    try:
        print("\n========== RECEBEU UPLOAD ==========", flush=True)
        print(f"Arquivo: {file.filename} | Tipo: {file.content_type}", flush=True)

        upload_id = str(uuid.uuid4())
        caminho = f"uploads/{upload_id}.jpg"

        with open(caminho, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        caminhos_temporarios.append(caminho)

        img = cv2.imread(caminho)

        if img is None:
            return {"erro": "Erro ao abrir imagem"}

        print("Criando versões...", flush=True)

        versoes = criar_versoes_imagem(caminho)
        caminhos_temporarios.extend(versoes)

        print("Iniciando OCR...", flush=True)

        texto = rodar_ocr_em_versoes(versoes)

        print("\n========== TEXTO OCR ==========", flush=True)
        print(texto, flush=True)

        dados_extraidos = extrair_dados_aveiro(texto)

        morada = formatar_morada_para_saida(dados_extraidos["morada"])
        codigo = dados_extraidos["codigo_postal"]
        cidade = dados_extraidos.get("cidade", "")

        for item in dados_extraidos.get("todos_resultados", []):
            if item.get("morada") and item.get("morada") != "Não encontrada":
                item["morada"] = formatar_morada_para_saida(item["morada"])

        uploads_pendentes[upload_id] = {
            "morada": morada,
            "codigo_postal": codigo,
            "cidade": cidade,
            "texto_ocr": texto,
            "geo_validada": dados_extraidos.get("geo_validada", False),
            "geo_motivo": dados_extraidos.get("geo_motivo", ""),
            "geo_display_name": dados_extraidos.get("geo_display_name", ""),
            "codigo_incompleto": dados_extraidos.get("codigo_incompleto", False),
            "todos_resultados": dados_extraidos["todos_resultados"],
        }

        print(
            f"\nID: {upload_id} | Morada: {morada} | CP: {codigo} | Cidade: {cidade}",
            flush=True,
        )

        return {
            "status": "aguardando_confirmacao",
            "mensagem": "Confirme ou edite os dados antes de adicionar ao lote.",
            "upload_id": upload_id,
            "morada": morada,
            "codigo_postal": codigo,
            "cidade": cidade,
            "geo_validada": dados_extraidos.get("geo_validada", False),
            "geo_motivo": dados_extraidos.get("geo_motivo", ""),
            "geo_display_name": dados_extraidos.get("geo_display_name", ""),
            "codigo_incompleto": dados_extraidos.get("codigo_incompleto", False),
            "texto_ocr": texto if texto else "Nenhum texto encontrado",
            "todos_resultados": dados_extraidos["todos_resultados"],
            "total_lote": len(lote_confirmado),
            "filtro": "Somente códigos postais 3800 e 3810 — moradas Portugal/Aveiro",
        }

    except Exception as e:
        print("\n========== ERRO ==========", flush=True)
        traceback.print_exc()

        return {"erro": str(e)}

    finally:
        for p in set(caminhos_temporarios):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass


# =========================
# CONFIRMAR NO LOTE
# =========================

@app.post("/confirmar")
async def confirmar(payload: ConfirmarPayload):
    try:
        morada = formatar_morada_para_saida(payload.morada)
        codigo = payload.codigo_postal.strip()
        cidade = limpar_linha(corrigir_ocr_para_morada(payload.cidade or ""))
        texto_ocr = payload.texto_ocr or ""

        if payload.upload_id and payload.upload_id in uploads_pendentes:
            pendente = uploads_pendentes[payload.upload_id]

            if not texto_ocr:
                texto_ocr = pendente.get("texto_ocr", "")

        # IMPORTANTE:
        # Aqui continua estrito. Não deixa adicionar/exportar só 3800 ou 3810.
        if not codigo_postal_valido(codigo):
            return {
                "erro": "Código postal incompleto ou inválido. Confirme o código postal completo no formato 3800-XXX ou 3810-XXX."
            }

        if not morada or morada == "Não encontrada":
            return {
                "erro": "Morada vazia. Confirme ou escreva a morada correta."
            }

        if eh_morada_nao_portugal(morada):
            return {
                "erro": "A morada indicada não parece ser portuguesa. Por favor, verifique."
            }

        lote_confirmado.append({
            "Morada": morada,
            "Código Postal": codigo,
            "Cidade": cidade,
            "Texto OCR": texto_ocr,
        })

        if payload.upload_id and payload.upload_id in uploads_pendentes:
            del uploads_pendentes[payload.upload_id]

        return {
            "status": "adicionado_ao_lote",
            "morada": morada,
            "codigo_postal": codigo,
            "cidade": cidade,
            "total_lote": len(lote_confirmado),
        }

    except Exception as e:
        print("\n========== ERRO AO CONFIRMAR ==========", flush=True)
        traceback.print_exc()

        return {"erro": str(e)}


# =========================
# LOTE
# =========================

@app.get("/resumo-lote")
async def resumo_lote():
    return {
        "total": len(lote_confirmado),
        "itens": lote_confirmado,
    }


@app.post("/limpar-lote")
async def limpar_lote():
    lote_confirmado.clear()
    uploads_pendentes.clear()
    limpar_pasta_exports()

    return {
        "status": "lote_limpo",
        "total": 0,
    }


# =========================
# DOWNLOAD
# =========================

@app.get("/download-excel")
async def download_excel():
    snapshot = list(lote_confirmado)

    export_id = uuid.uuid4().hex
    excel_path = f"exports/resultado_{export_id}.xlsx"
    csv_path = f"exports/resultado_{export_id}.csv"

    salvar_lote_em_arquivos(
        snapshot=snapshot,
        excel_path=excel_path,
        csv_path=csv_path,
    )

    return FileResponse(
        path=excel_path,
        filename="resultado.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(
            limpar_ficheiros_exportacao,
            excel_path,
            csv_path,
        ),
    )


@app.get("/download-csv")
async def download_csv():
    snapshot = list(lote_confirmado)

    export_id = uuid.uuid4().hex
    excel_path = f"exports/resultado_{export_id}.xlsx"
    csv_path = f"exports/resultado_{export_id}.csv"

    salvar_lote_em_arquivos(
        snapshot=snapshot,
        excel_path=excel_path,
        csv_path=csv_path,
    )

    return FileResponse(
        path=csv_path,
        filename="resultado.csv",
        media_type="text/csv",
        background=BackgroundTask(
            limpar_ficheiros_exportacao,
            excel_path,
            csv_path,
        ),
    )
